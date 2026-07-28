from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
import pandas as pd
import re
import os
import uuid
import json
import hashlib
import hmac
import io
import secrets
import time
import requests
from datetime import datetime, timedelta
from crm import (
    CALL_DIRECTIONS,
    CALL_OUTCOMES,
    CRM_PRIORITIES,
    CRM_STATUSES,
    EVIDENCE_CONFIDENCE,
    EVIDENCE_OUTCOMES,
    EVIDENCE_TYPES,
    IDENTITY_MATCHES,
    PIPELINE_STAGES,
    RESEARCH_STATUSES,
    CRMRepository,
)
from qualification import qualify_leads
from assessor import AssessorResult, TulsaAssessorClient, normalize_account_no, verification_decision

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_bytes(32)
app.config['APP_LOGIN_EMAIL'] = os.environ.get('APP_LOGIN_EMAIL', '').strip().lower()
app.config['APP_LOGIN_PASSWORD'] = os.environ.get('APP_LOGIN_PASSWORD', '')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('RENDER'))
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.config['DATABASE_URL'] = os.environ.get('DATABASE_URL')
app.config['ASSESSOR_BATCH_LIMIT'] = 25
app.config['ASSESSOR_CLIENT_FACTORY'] = TulsaAssessorClient
app.config['CRM_DATABASE'] = os.environ.get(
    'CRM_DATABASE',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'leadcleaner.db')
)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
LOGIN_ATTEMPTS = {}
LOGIN_WINDOW_SECONDS = 15 * 60
LOGIN_MAX_ATTEMPTS = 5


def authentication_configured():
    return bool(
        app.config.get('APP_LOGIN_EMAIL')
        and app.config.get('APP_LOGIN_PASSWORD')
    )


def csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_urlsafe(32)
    return session['csrf_token']


@app.context_processor
def inject_security_context():
    return {'csrf_token': csrf_token()}


@app.before_request
def protect_private_workspace():
    if app.config.get('TESTING') and not app.config.get('TEST_AUTH_ENABLED'):
        return None
    if request.endpoint in ('login', 'api_health') or request.endpoint == 'static':
        return None
    if not session.get('authenticated'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Authentication required'}), 401
        return redirect(url_for('login', next=request.path))
    if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
        supplied = (
            request.headers.get('X-CSRF-Token', '')
            or request.form.get('csrf_token', '')
        )
        expected = session.get('csrf_token', '')
        if not expected or not hmac.compare_digest(supplied, expected):
            return jsonify({'error': 'Security token is missing or invalid'}), 403
    return None


@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['Permissions-Policy'] = (
        'camera=(), microphone=(), geolocation=()'
    )
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    if request.endpoint != 'api_health':
        response.headers['Cache-Control'] = 'no-store, private'
    if app.config.get('SESSION_COOKIE_SECURE'):
        response.headers['Strict-Transport-Security'] = (
            'max-age=31536000; includeSubDomains'
        )
    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    configured = authentication_configured()
    if request.method == 'POST':
        attempt_key = request.remote_addr or 'unknown'
        now = time.monotonic()
        recent_attempts = [
            timestamp
            for timestamp in LOGIN_ATTEMPTS.get(attempt_key, [])
            if now - timestamp < LOGIN_WINDOW_SECONDS
        ]
        LOGIN_ATTEMPTS[attempt_key] = recent_attempts
        supplied_token = request.form.get('csrf_token', '')
        expected_token = session.get('csrf_token', '')
        if len(recent_attempts) >= LOGIN_MAX_ATTEMPTS:
            error = 'Too many sign-in attempts. Please wait 15 minutes.'
        elif not expected_token or not hmac.compare_digest(
            supplied_token,
            expected_token,
        ):
            error = 'Your secure session expired. Please try again.'
        elif not configured:
            error = 'Private access has not been configured on the server.'
        else:
            email_matches = hmac.compare_digest(
                request.form.get('email', '').strip().lower(),
                app.config['APP_LOGIN_EMAIL'],
            )
            password_matches = hmac.compare_digest(
                request.form.get('password', ''),
                app.config['APP_LOGIN_PASSWORD'],
            )
            if email_matches and password_matches:
                LOGIN_ATTEMPTS.pop(attempt_key, None)
                session.clear()
                session['authenticated'] = True
                session['user_email'] = app.config['APP_LOGIN_EMAIL']
                session['csrf_token'] = secrets.token_urlsafe(32)
                session.permanent = True
                destination = request.args.get('next', '/')
                if not destination.startswith('/') or destination.startswith('//'):
                    destination = '/'
                return redirect(destination)
            LOGIN_ATTEMPTS[attempt_key].append(now)
            error = 'The email or password is incorrect.'
    return render_template(
        'login.html',
        error=error,
        configured=configured,
    )


@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return redirect(url_for('login'))


def get_crm():
    database_target = app.config.get('DATABASE_URL') or app.config['CRM_DATABASE']
    repository = CRMRepository(database_target)
    repository.initialize()
    return repository


def save_job_meta(meta):
    meta_path = os.path.join(
        app.config['OUTPUT_FOLDER'],
        f"{meta['uid']}_meta.json",
    )
    with open(meta_path, 'w') as file_handle:
        json.dump(meta, file_handle)
    get_crm().save_processing_job(meta)
    return meta_path


def load_job_meta(job_id):
    meta_path = os.path.join(
        app.config['OUTPUT_FOLDER'],
        f'{job_id}_meta.json',
    )
    if os.path.exists(meta_path):
        with open(meta_path) as file_handle:
            return json.load(file_handle)
    meta = get_crm().get_processing_job(job_id)
    if not meta:
        return None
    meta.pop('_created_at', None)
    meta.pop('_updated_at', None)
    with open(meta_path, 'w') as file_handle:
        json.dump(meta, file_handle)
    return meta


def persist_artifact(job_id, kind, filename, path):
    with open(path, 'rb') as file_handle:
        content = file_handle.read()
    return get_crm().save_processing_artifact(
        job_id,
        kind,
        filename,
        content,
    )


def materialize_artifact(job_id, kind, filename, folder):
    safe_name = os.path.basename(filename)
    path = os.path.join(folder, safe_name)
    if os.path.exists(path):
        return path
    artifact = get_crm().get_processing_artifact(
        job_id=job_id,
        kind=kind,
    )
    if not artifact:
        return None
    with open(path, 'wb') as file_handle:
        file_handle.write(artifact['content'])
    return path


JOB_STATUS_LABELS = {
    'uploaded': 'Upload received',
    'qualifying': 'Qualifying records',
    'qualification_ready': 'Ready for Assessor',
    'assessor_in_progress': 'Assessor in progress',
    'ready_for_approval': 'Ready for approval',
    'imported': 'Imported to CRM',
    'failed': 'Needs attention',
}


def processing_job_snapshot(meta):
    status = meta.get('status', 'qualification_ready')
    stats = meta.get('stats', {})
    assessor = meta.get('assessor_progress', {})
    total = int(
        assessor.get('total')
        or stats.get('prequalified')
        or stats.get('final')
        or 0
    )
    checked = int(assessor.get('checked', 0))
    if status == 'uploaded':
        progress = 10
    elif status == 'qualifying':
        progress = 35
    elif status == 'qualification_ready':
        progress = 55
    elif status == 'assessor_in_progress':
        progress = 55 + round(35 * checked / max(total, 1))
    elif status == 'ready_for_approval':
        progress = 90
    elif status == 'imported':
        progress = 100
    else:
        progress = int(meta.get('progress', 0))
    output_filename = meta.get(
        'assessor_output_filename',
        meta.get('output_filename'),
    )
    return {
        'id': meta.get('uid'),
        'status': status,
        'status_label': JOB_STATUS_LABELS.get(status, status.replace('_', ' ').title()),
        'progress': min(max(progress, 0), 100),
        'tax_year': meta.get('tax_year'),
        'source_filename': meta.get('source_filename'),
        'output_filename': output_filename,
        'stats': stats,
        'assessor': {
            'checked': checked,
            'total': total,
            'remaining': max(total - checked, 0),
            'decision_counts': assessor.get('decision_counts', {}),
        },
        'error': meta.get('error'),
        'actions': {
            'verify_assessor': status in (
                'qualification_ready',
                'assessor_in_progress',
            ),
            'preview_approval': status in (
                'assessor_in_progress',
                'ready_for_approval',
                'imported',
            ) and bool(meta.get('assessor_output_filename')),
            'download': bool(output_filename),
        },
    }

# ─── SKIP TRACING CONFIG ──────────────────────────────────────────────────────
SKIP_TRACE_PROVIDER = 'none'
SKIP_TRACE_API_KEY = os.environ.get('SKIP_TRACE_API_KEY', '')

# ─── BUSINESS PATTERNS ───────────────────────────────────────────────────────

BUSINESS_PATTERNS = [
    r'\bLLC\b', r'\bL\.L\.C\.?\b',
    r'\bINC\.?\b', r'\bINCORPORATED\b',
    r'\bCORP\.?\b', r'\bCORPORATION\b',
    r'\bLTD\.?\b', r'\bLIMITED\b',
    r'\bCOMPANY\b', r'\bCO\.\b',
    r'\bGROUP\b', r'\bENTERPRISES?\b',
    r'\bASSOCIATES?\b', r'\bPARTNERS?\b',
    r'\bHOLDINGS?\b', r'\bREALTY\b',
    r'\bREAL ESTATE\b', r'\bPROPERTIES\b',
    r'\bINVESTMENTS?\b', r'\bVENTURES?\b',
    r'\bCHURCH\b', r'\bCHAPEL\b', r'\bMINISTRIES?\b',
    r'\bFOUNDATION\b', r'\bCATTLE\b', r'\bRANCH\b',
    r'\bFARMS?\b', r'\bCULTIVATION\b',
    r'\bL\.?\s*P\.?\b',    # Limited Partnership (ex: 'OKLAHOMA L P')
    r'\bATTY\b',            # escritórios de advocacia abreviados
]

CANNABIS_PATTERNS = [
    r'\bCANNABIS\b',
    r'\bDISPENSARY\b',
    r'\bDISPENSARIES\b',
    r'\bMARIJUANA\b',
    r'\bMARIHUANA\b',
    r'\bHEMP\b',
    r'\bCBD\b',
    r'\bTHC\b',
    r'\bMMJ\b',
    r'\bMEDICINAL\b',
    r'\b420\b',
    r'\bGANJA\b',
    r'\bWEED\s+(CO|LLC|INC|CORP|GROUP)\b',
    # Termos de gíria de dispensárias (ex: "MMA SKUNK GROW", "MMA BIG BUDS")
    # — 'MMA' sozinho é ambíguo (pode ser nome/sigla legítima), por isso só
    # conta como cannabis quando aparece junto de gíria específica do ramo.
    r'\bMMA\b.*\b(SKUNK|GROW|GREEN\s*MEDS?|KUSH|BUDS?|DANK|OG)\b',
    r'\b(SKUNK|KUSH|DANK)\b',   # baixo risco de serem apelidos reais
    r'\bBUDS\b',                 # plural — 'Bud' sozinho é nickname comum, 'Buds' não
]

DECEASED_PATTERNS = [
    r'\bDECEASED\b',
    r'\bESTATE\s+OF\b',
    r'(?<!REAL\s)\bESTATE\b(?!\s+LLC)(?!\s+TRUST)(?!\s+SERIES)',
    r'\bHEIRS?\s+OF\b',
    r'\bPR\s+OF\s+THE\s+ESTATE\b',
    r'\bPERSONAL\s+REP(RESENTATIVE)?\b',
    r'\bEXECUTOR\b',
    r'\bSURVIVING\s+(SPOUSE|HEIR)\b',
    r'\bIN\s+CARE\s+OF\s+ESTATE\b',
    r'\bC/?O\s+ESTATE\b',
]


def find_column(df, keywords):
    """Find the first column whose name contains ALL given keywords (case-insensitive,
    ignoring spaces). Handles variants like 'Owner Name', 'PROPERTY OWNER NAME',
    'OwnerName', 'TotalDue', 'Total Due', etc."""
    for col in df.columns:
        normalized = re.sub(r'[^A-Z0-9]', '', str(col).upper())
        for kw in keywords:
            kw_norm = re.sub(r'[^A-Z0-9]', '', kw.upper())
            if not kw_norm or kw_norm not in normalized:
                break
        else:
            return col
    return None


def is_business(name):
    if pd.isna(name):
        return False
    name_upper = str(name).upper()
    if re.search(r'\bTRUST(EE)?\b', name_upper):
        return False
    for pattern in BUSINESS_PATTERNS:
        if re.search(pattern, name_upper):
            return True
    return False


def is_cannabis(name):
    if pd.isna(name):
        return False
    name_upper = str(name).upper()
    for pattern in CANNABIS_PATTERNS:
        if re.search(pattern, name_upper):
            return True
    return False


def is_likely_deceased(name, comments=None):
    fields_to_check = []
    if not pd.isna(name):
        fields_to_check.append(str(name).upper().strip())
    if comments is not None and not pd.isna(comments):
        fields_to_check.append(str(comments).upper())

    for field in fields_to_check:
        if re.search(r'\bREAL\s+ESTATE\b', field):
            continue
        for pattern in DECEASED_PATTERNS:
            if re.search(pattern, field):
                return True
    return False


def reorder_columns_for_readability(df, owner_col, total_due_col):
    """Push the columns a real-estate investor actually looks at first
    (owner, deceased flag, amount owed, phone, address), and push
    technical/GIS columns (PID, legal description, SecTwnRng, etc.) to the end."""
    priority_names = []
    for candidate in [
        owner_col,
        'Deceased Owner (Flagged)',
        total_due_col,
        find_column(df, ['PHONE']),
        find_column(df, ['ADDRESS']) or find_column(df, ['ST_NO']),
        find_column(df, ['ST_NAME']),
        find_column(df, ['ST_STREET', 'TYPE']),
        find_column(df, ['OWNR_ADDR', '6']),
        find_column(df, ['OWNR_ADDR', 'ST']),
        find_column(df, ['ZIP']),
    ]:
        if candidate and candidate in df.columns and candidate not in priority_names:
            priority_names.append(candidate)

    remaining = [c for c in df.columns if c not in priority_names]
    return df[priority_names + remaining]


def save_excel_formatted(sheets: dict, output_path):
    """Write a dict of {sheet_name: DataFrame} to xlsx with a clean,
    readable look: bold header row, frozen header, auto-sized columns."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            header_fill = PatternFill(start_color='1F1B16', end_color='1F1B16', fill_type='solid')
            header_font = Font(bold=True, color='D4AF37')
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical='center')

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            for i, col in enumerate(sheet_df.columns, start=1):
                max_len = max(
                    [len(str(col))] + [len(str(v)) for v in sheet_df[col].astype(str).head(500)]
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)

            ws.row_dimensions[1].height = 20


def compute_absentee_signal(df, mail_addr_col, mail_city_col, prop_city_col):
    """
    Filtro 1 — sinal fraco e gratuito de 'proprietário provavelmente ausente
    ou falecido', baseado só nos dados que o condado já fornece:
      - morada de correspondência contém 'C/O' ou 'PO BOX'
      - cidade de correspondência difere da cidade do imóvel

    Devolve uma Series com: '' (sem sinal), 'Weak' (um dos dois sinais),
    ou 'Strong' (os dois sinais ao mesmo tempo — muito mais fiável).

    Isto NÃO confirma óbito — é só um filtro de prioridade para reduzir
    o volume que precisa de verificação mais cara (OK2Explore, OSCN, etc.)
    """
    n = len(df)
    if not (mail_addr_col and mail_city_col and prop_city_col):
        return pd.Series([''] * n, index=df.index)

    addr = df[mail_addr_col].fillna('').astype(str).str.upper()
    co_po = addr.str.contains(r'\bC/?O\b|\bP\.?O\.?\s*BOX\b', regex=True)

    prop_city = (
        df[prop_city_col].fillna('').astype(str).str.upper()
        .str.replace('CITY OF ', '', regex=False)
        .str.replace(' COUNTY', '', regex=False)
        .str.strip()
    )
    mail_city = df[mail_city_col].fillna('').astype(str).str.upper().str.strip()
    mismatch = (prop_city != '') & (mail_city != '') & (prop_city != mail_city)

    strength = pd.Series([''] * n, index=df.index)
    strength[co_po | mismatch] = 'Weak'
    strength[co_po & mismatch] = 'Strong'
    return strength


def clean_leads(df, tax_year):
    stats = {'original': len(df)}

    # ── Resolve real column names dynamically (fixes 'Owner Name' vs
    #    'PROPERTY OWNER NAME', 'TotalDue' vs 'Total Due', etc.) ──
    owner_col = find_column(df, ['OWNER', 'NAME'])
    if owner_col is None:
        raise ValueError(
            f"Coluna do nome do proprietario nao encontrada. "
            f"Colunas disponiveis: {list(df.columns)}"
        )

    total_due_col = find_column(df, ['TOTAL', 'DUE'])
    comments_col = find_column(df, ['COMMENT'])
    tax_col = find_column(df, ['TAX', 'YEAR'])
    mail_addr_col = find_column(df, ['ADDRESS'])
    mail_city_col = find_column(df, ['OWNR_ADDR', '6'])
    prop_city_col = find_column(df, ['ST_CITY'])

    # Detect tax year column (only filters if the column actually exists)
    if tax_col:
        df[tax_col] = pd.to_numeric(df[tax_col], errors='coerce')
        df = df[df[tax_col] == tax_year].copy()
    stats['after_year_filter'] = len(df)

    # Flag deceased BEFORE removing businesses (estate records are valuable leads)
    df['_deceased'] = df.apply(
        lambda row: is_likely_deceased(
            row.get(owner_col, ''),
            row.get(comments_col, None) if comments_col else None
        ),
        axis=1
    )

    # Remove cannabis businesses
    df['_is_cannabis'] = df[owner_col].apply(is_cannabis)
    stats['removed_cannabis'] = int(df['_is_cannabis'].sum())
    df = df[~df['_is_cannabis']].copy()

    # Remove other businesses (keep trusts)
    df['_is_business'] = df[owner_col].apply(is_business)
    df = df[~df['_is_business']].copy()
    df = df.drop(columns=['_is_cannabis', '_is_business'])

    stats['after_business_filter'] = len(df)

    # Add Deceased Owner column (kept in main sheet too — valuable leads)
    df['Deceased Owner (Flagged)'] = df['_deceased'].map(
        {True: 'YES - Verify', False: ''}
    )
    stats['deceased_flagged'] = int(df['_deceased'].sum())

    # Filtro 1 — sinal fraco de morada suspeita (C/O, PO BOX, cidade divergente)
    df['_absentee_signal'] = compute_absentee_signal(df, mail_addr_col, mail_city_col, prop_city_col)
    df['Absentee/Suspicious Mailing (Verify)'] = df['_absentee_signal']
    stats['absentee_signal_strong'] = int((df['_absentee_signal'] == 'Strong').sum())
    stats['absentee_signal_weak'] = int((df['_absentee_signal'] == 'Weak').sum())

    # Build the separate "deceased owners" tab Daryl asked for (regex-confirmed)
    deceased_df = df[df['_deceased']].drop(columns=['_deceased', '_absentee_signal']).copy()

    # Build the "suspected — verify manually" tab: heuristic-flagged owners
    # who are NOT already in the regex-confirmed deceased tab (avoid duplicates),
    # ordered so 'Strong' (both signals) comes before 'Weak' (one signal only)
    has_signal = (df['_absentee_signal'] != '') & ~df['_deceased']
    suspected_df = df[has_signal].drop(columns=['_deceased', '_absentee_signal']).copy()
    if len(suspected_df):
        strength_order = {'Strong': 0, 'Weak': 1}
        suspected_df['_sort'] = suspected_df['Absentee/Suspicious Mailing (Verify)'].map(strength_order)
        suspected_df = suspected_df.sort_values('_sort').drop(columns=['_sort'])

    df = df.drop(columns=['_deceased', '_absentee_signal'])

    # Remove completely empty rows
    df = df.dropna(how='all')
    stats['final'] = len(df)

    # Sort by Total Due descending (works regardless of 'Total Due' vs 'TotalDue')
    if total_due_col:
        df = df.sort_values(total_due_col, ascending=False)
        deceased_df = deceased_df.sort_values(total_due_col, ascending=False) if len(deceased_df) else deceased_df
        if len(suspected_df):
            strength_order = {'Strong': 0, 'Weak': 1}
            suspected_df['_sort'] = suspected_df['Absentee/Suspicious Mailing (Verify)'].map(strength_order)
            suspected_df = suspected_df.sort_values(['_sort', total_due_col], ascending=[True, False]).drop(columns=['_sort'])

    stats['removed_year']     = stats['original'] - stats['after_year_filter']
    stats['removed_business'] = stats['after_year_filter'] - stats['after_business_filter']
    phone_col = find_column(df, ['PHONE'])
    stats['with_phone']       = int(df[phone_col].notna().sum()) if phone_col else 0
    stats['without_phone']    = stats['final'] - stats['with_phone']

    stats = {k: int(v) for k, v in stats.items()}

    # Reorder columns so the useful ones (owner, deceased flag, amount owed,
    # phone, address) come first and technical/GIS columns come last
    df = reorder_columns_for_readability(df, owner_col, total_due_col)
    if len(deceased_df):
        deceased_df = reorder_columns_for_readability(deceased_df, owner_col, total_due_col)
    if len(suspected_df):
        suspected_df = reorder_columns_for_readability(suspected_df, owner_col, total_due_col)

    return df, deceased_df, suspected_df, stats


# ─── SKIP TRACING ─────────────────────────────────────────────────────────────

def run_skip_tracing(df):
    if SKIP_TRACE_PROVIDER == 'none' or not SKIP_TRACE_API_KEY:
        return df, {'error': 'No skip tracing provider configured.'}
    elif SKIP_TRACE_PROVIDER == 'batchdata':
        return _skip_trace_batchdata(df)
    elif SKIP_TRACE_PROVIDER == 'tracerfy':
        return _skip_trace_tracerfy(df)
    else:
        return df, {'error': f'Unknown provider: {SKIP_TRACE_PROVIDER}'}


def _skip_trace_batchdata(df):
    try:
        records = []
        for _, row in df.iterrows():
            record = {
                'address':   str(row.get('ST_NO', '')) + ' ' + str(row.get('ST_NAME', '')),
                'city':      str(row.get('OWNR_ADDR 6', '')),
                'state':     str(row.get('OWNR_ADDR ST', '')),
                'zip':       str(row.get('ZIP', '')),
                'firstName': '',
                'lastName':  str(row.get('Owner Name', '')),
            }
            records.append(record)

        response = requests.post(
            'https://api.batchdata.com/api/v1/property/skip-trace',
            headers={
                'Authorization': f'Bearer {SKIP_TRACE_API_KEY}',
                'Content-Type': 'application/json'
            },
            json={'requests': records},
            timeout=120
        )
        response.raise_for_status()
        results = response.json()

        phones = []
        for result in results.get('results', []):
            contacts = result.get('results', {}).get('phoneNumbers', [])
            phone_list = [p.get('number', '') for p in contacts[:8]]
            while len(phone_list) < 8:
                phone_list.append('')
            phones.append(phone_list)

        for i in range(1, 9):
            df[f'Phone {i}'] = [p[i-1] if i-1 < len(p) else '' for p in phones]

        stats = {
            'provider': 'BatchData',
            'records_sent': len(records),
            'records_matched': sum(1 for p in phones if any(p))
        }
        return df, stats

    except Exception as e:
        return df, {'error': str(e)}


def _skip_trace_tracerfy(df):
    try:
        import csv, io, time

        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['address', 'city', 'state', 'zip', 'first_name', 'last_name'])
        for _, row in df.iterrows():
            address = f"{row.get('ST_NO', '')} {row.get('ST_NAME', '')}".strip()
            writer.writerow([
                address,
                row.get('OWNR_ADDR 6', ''),
                row.get('OWNR_ADDR ST', ''),
                row.get('ZIP', ''),
                '',
                row.get('Owner Name', '')
            ])
        csv_content = csv_buffer.getvalue()

        response = requests.post(
            'https://api.tracerfy.com/trace/',
            headers={'Authorization': f'Bearer {SKIP_TRACE_API_KEY}'},
            files={'file': ('leads.csv', csv_content, 'text/csv')},
            data={'trace_type': 'normal'},
            timeout=60
        )
        response.raise_for_status()
        queue_id = response.json().get('queue_id')

        download_url = None
        for _ in range(30):
            time.sleep(10)
            status_resp = requests.get(
                f'https://api.tracerfy.com/queue/{queue_id}/',
                headers={'Authorization': f'Bearer {SKIP_TRACE_API_KEY}'}
            )
            status_data = status_resp.json()
            if not status_data.get('pending'):
                download_url = status_data.get('download_url')
                break

        result_resp = requests.get(download_url)
        result_df = pd.read_csv(io.StringIO(result_resp.text))

        phone_cols = [c for c in result_df.columns if 'phone' in c.lower()]
        for col in phone_cols[:8]:
            df[col] = result_df[col].values if len(result_df) == len(df) else ''

        stats = {
            'provider': 'Tracerfy',
            'records_sent': len(df),
            'records_matched': result_df[phone_cols[0]].notna().sum() if phone_cols else 0
        }
        return df, stats

    except Exception as e:
        return df, {'error': str(e)}


# ─── ROUTES ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


def get_dashboard_snapshot():
    """Build the dashboard from completed processing jobs.

    PostgreSQL is the durable source of truth. The local output directory is
    only a rebuildable cache for generated workbooks.
    """
    jobs = []
    repository = get_crm()
    stored_jobs = repository.list_processing_jobs(limit=20)
    stored_ids = {meta.get('uid') for meta in stored_jobs}
    # Preserve visibility of jobs created before durable storage was deployed.
    # New jobs always enter PostgreSQL; these local records are a compatibility
    # bridge and disappear naturally when Render replaces the old filesystem.
    for filename in os.listdir(app.config['OUTPUT_FOLDER']):
        if not filename.endswith('_meta.json'):
            continue
        meta_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        try:
            with open(meta_path, encoding='utf-8') as meta_file:
                meta = json.load(meta_file)
        except (OSError, ValueError):
            continue
        if meta.get('uid') in stored_ids:
            continue
        meta['_updated_at'] = datetime.fromtimestamp(
            os.path.getmtime(meta_path)
        ).isoformat()
        stored_jobs.append(meta)

    for meta in stored_jobs:
        output_filename = meta.get(
            'assessor_output_filename',
            meta.get('output_filename', ''),
        )
        completed_value = (
            meta.get('_updated_at')
            or meta.get('created_at')
            or meta.get('_created_at')
        )
        try:
            completed_at = datetime.fromisoformat(completed_value)
        except (TypeError, ValueError):
            completed_at = datetime.now()
        artifact = (
            repository.get_processing_artifact(filename=output_filename)
            if output_filename
            else None
        )
        local_output = os.path.join(
            app.config['OUTPUT_FOLDER'],
            os.path.basename(output_filename),
        )
        stats = meta.get('stats', {})
        workflow = processing_job_snapshot(meta)
        jobs.append({
            'id': meta.get('uid'),
            'status': workflow['status'],
            'status_label': workflow['status_label'],
            'progress': workflow['progress'],
            'tax_year': meta.get('tax_year'),
            'completed_at': completed_at.isoformat(),
            'completed_label': completed_at.strftime('%b %d, %Y · %I:%M %p'),
            'output_filename': output_filename,
            'download_available': bool(
                artifact or (output_filename and os.path.exists(local_output))
            ),
            'stats': stats,
        })

    jobs.sort(key=lambda job: job['completed_at'], reverse=True)
    latest = jobs[0] if jobs else None
    latest_stats = latest['stats'] if latest else {}

    crm_metrics = get_crm().dashboard_metrics()
    has_crm_data = crm_metrics['actionable_leads'] > 0
    metrics = crm_metrics if has_crm_data else {
        'actionable_leads': int(latest_stats.get('final', 0)),
        'deceased_signals': int(latest_stats.get('deceased_flagged', 0)),
        'research_queue': int(
            latest_stats.get(
                'review',
                latest_stats.get('absentee_signal_strong', 0)
                + latest_stats.get('absentee_signal_weak', 0),
            )
        ),
        'contacts_found': int(latest_stats.get('with_phone', 0)),
        'overdue_follow_ups': 0,
        'follow_ups_due': 0,
    }

    return {
        'has_data': bool(latest) or has_crm_data,
        'generated_at': datetime.now().isoformat(),
        'metrics': metrics,
        'attention': [
            {
                'type': 'follow_up',
                'priority': 'High',
                'title': 'Complete today’s follow-ups',
                'detail': f"{metrics.get('follow_ups_due', 0)} owners are due for contact",
                'count': metrics.get('follow_ups_due', 0),
            },
            {
                'type': 'deceased',
                'priority': 'High',
                'title': 'Review deceased-owner evidence',
                'detail': f"{metrics['deceased_signals']} records flagged by county-data patterns",
                'count': metrics['deceased_signals'],
            },
            {
                'type': 'research',
                'priority': 'Medium',
                'title': 'Investigate suspicious mailing records',
                'detail': f"{metrics['research_queue']} records still require evidence review",
                'count': metrics['research_queue'],
            },
            {
                'type': 'contact',
                'priority': 'Normal',
                'title': 'Prepare leads for contact',
                'detail': (
                    f"{max(metrics['actionable_leads'] - metrics['contacts_found'], 0)} "
                    "records still need contact data"
                ),
                'count': max(
                    metrics['actionable_leads'] - metrics['contacts_found'], 0
                ),
            },
        ] if latest or has_crm_data else [],
        'latest_job': latest,
        'recent_jobs': jobs[:5],
    }


@app.route('/api/dashboard')
def dashboard_snapshot():
    return jsonify(get_dashboard_snapshot())


@app.route('/leads')
def leads_page():
    return render_template(
        'leads.html',
        page_mode='leads',
        statuses=CRM_STATUSES,
        priorities=CRM_PRIORITIES,
        research_statuses=RESEARCH_STATUSES,
        evidence_types=EVIDENCE_TYPES,
        evidence_outcomes=EVIDENCE_OUTCOMES,
        evidence_confidence=EVIDENCE_CONFIDENCE,
        identity_matches=IDENTITY_MATCHES,
        call_outcomes=CALL_OUTCOMES,
        call_directions=CALL_DIRECTIONS,
    )


@app.route('/today')
def today_page():
    return render_template(
        'leads.html',
        page_mode='today',
        statuses=CRM_STATUSES,
        priorities=CRM_PRIORITIES,
        research_statuses=RESEARCH_STATUSES,
        evidence_types=EVIDENCE_TYPES,
        evidence_outcomes=EVIDENCE_OUTCOMES,
        evidence_confidence=EVIDENCE_CONFIDENCE,
        identity_matches=IDENTITY_MATCHES,
        call_outcomes=CALL_OUTCOMES,
        call_directions=CALL_DIRECTIONS,
    )


@app.route('/pipeline')
def pipeline_page():
    return render_template(
        'pipeline.html',
        pipeline_stages=PIPELINE_STAGES,
    )


@app.route('/properties')
def properties_page():
    return render_template(
        'properties.html',
        statuses=CRM_STATUSES,
    )


@app.route('/research')
def research_page():
    return render_template(
        'leads.html',
        page_mode='research',
        statuses=CRM_STATUSES,
        priorities=CRM_PRIORITIES,
        research_statuses=RESEARCH_STATUSES,
        evidence_types=EVIDENCE_TYPES,
        evidence_outcomes=EVIDENCE_OUTCOMES,
        evidence_confidence=EVIDENCE_CONFIDENCE,
        identity_matches=IDENTITY_MATCHES,
        call_outcomes=CALL_OUTCOMES,
        call_directions=CALL_DIRECTIONS,
    )


@app.route('/api/leads')
def api_leads():
    try:
        result = get_crm().list_leads(
            search=request.args.get('q', '').strip(),
            status=request.args.get('status', '').strip(),
            priority=request.args.get('priority', '').strip(),
            research_only=request.args.get('research_only', '').lower() == 'true',
            follow_up=request.args.get('follow_up', '').strip(),
            page=request.args.get('page', 1),
            per_page=request.args.get('per_page', 50),
        )
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid pagination or filter value'}), 400
    return jsonify(result)


@app.route('/api/pipeline')
def api_pipeline():
    try:
        board = get_crm().pipeline_board(
            cards_per_stage=request.args.get('cards_per_stage', 50)
        )
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid pipeline limit'}), 400
    return jsonify(board)


@app.route('/api/properties')
def api_properties():
    try:
        result = get_crm().list_properties(
            search=request.args.get('q', '').strip(),
            status=request.args.get('status', '').strip(),
            page=request.args.get('page', 1),
            per_page=request.args.get('per_page', 24),
        )
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid pagination or filter value'}), 400
    return jsonify(result)


@app.route('/api/health')
def api_health():
    return jsonify(get_crm().health())


@app.route('/api/leads/<int:lead_id>')
def api_lead_detail(lead_id):
    lead = get_crm().get_lead(lead_id)
    if not lead:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify(lead)


@app.route('/api/leads/<int:lead_id>', methods=['PATCH'])
def api_update_lead(lead_id):
    payload = request.get_json(silent=True) or {}
    try:
        lead = get_crm().update_lead(lead_id, payload)
    except ValueError as error:
        return jsonify({'error': str(error)}), 400
    if not lead:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify({'success': True, 'lead': lead})


@app.route('/api/leads/<int:lead_id>/notes', methods=['POST'])
def api_add_lead_note(lead_id):
    payload = request.get_json(silent=True) or {}
    try:
        note = get_crm().add_note(lead_id, payload.get('body'))
    except ValueError as error:
        return jsonify({'error': str(error)}), 400
    if not note:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify({'success': True, 'note': note}), 201


@app.route('/api/leads/<int:lead_id>/calls', methods=['POST'])
def api_log_lead_call(lead_id):
    payload = request.get_json(silent=True) or {}
    try:
        result = get_crm().log_call(lead_id, payload)
    except ValueError as error:
        return jsonify({'error': str(error)}), 400
    if not result:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify({'success': True, **result}), 201


@app.route('/api/leads/<int:lead_id>/evidence', methods=['POST'])
def api_add_lead_evidence(lead_id):
    payload = request.get_json(silent=True) or {}
    try:
        result = get_crm().add_evidence(lead_id, payload)
    except ValueError as error:
        return jsonify({'error': str(error)}), 400
    if not result:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify({'success': True, **result}), 201


@app.route(
    '/api/leads/<int:lead_id>/evidence/<int:evidence_id>',
    methods=['DELETE'],
)
def api_retract_lead_evidence(lead_id, evidence_id):
    payload = request.get_json(silent=True) or {}
    try:
        lead = get_crm().retract_evidence(
            lead_id,
            evidence_id,
            payload.get('reason'),
        )
    except ValueError as error:
        return jsonify({'error': str(error)}), 400
    if not lead:
        return jsonify({'error': 'Lead or evidence not found'}), 404
    return jsonify({'success': True, 'lead': lead})


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    tax_year = request.form.get('tax_year', '2023')
    try:
        tax_year = int(tax_year)
    except ValueError:
        return jsonify({'error': 'Invalid tax year'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv']:
        return jsonify({'error': 'File must be .xlsx, .xls or .csv'}), 400

    uid = str(uuid.uuid4())[:8]
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{uid}_input{ext}')
    file.save(upload_path)
    job_meta = {
        'uid': uid,
        'source_filename': file.filename,
        'tax_year': tax_year,
        'stats': {},
        'status': 'uploaded',
        'created_at': datetime.now().isoformat(),
    }
    save_job_meta(job_meta)
    persist_artifact(uid, 'source', os.path.basename(upload_path), upload_path)

    try:
        job_meta['status'] = 'qualifying'
        save_job_meta(job_meta)
        df = pd.read_csv(upload_path) if ext == '.csv' else pd.read_excel(upload_path, engine='openpyxl')
    except Exception as e:
        job_meta['status'] = 'failed'
        job_meta['error'] = f'Could not read file: {str(e)}'
        save_job_meta(job_meta)
        return jsonify({'error': f'Could not read file: {str(e)}'}), 400

    try:
        qualification = qualify_leads(df, tax_year)
        cleaned_df = qualification['qualified']
        stats = qualification['stats']
    except Exception as e:
        job_meta['status'] = 'failed'
        job_meta['error'] = f'Error during cleaning: {str(e)}'
        job_meta['columns_found'] = list(df.columns)
        save_job_meta(job_meta)
        return jsonify({'error': f'Error during cleaning: {str(e)}', 'columns_found': list(df.columns)}), 500

    date_str = datetime.now().strftime('%Y%m%d')
    output_filename = f'Clean_Leads_{tax_year}_{date_str}_{uid}.xlsx'
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

    save_excel_formatted(
        {
            'Prequalified - Verify': qualification['qualified'],
            'Needs Review': qualification['review'],
            'Deceased Research': qualification['deceased'],
            'Absentee Opportunities': qualification['absentee'],
            'Excluded Records': qualification['excluded'],
        },
        output_path
    )

    job_meta['output_filename'] = output_filename
    job_meta['stats'] = stats
    job_meta['status'] = 'qualification_ready'
    job_meta.pop('error', None)
    save_job_meta(job_meta)
    persist_artifact(uid, 'qualification', output_filename, output_path)

    # Qualification runs remain reviewable exports. A separate explicit commit
    # action will import approved leads into the production CRM.
    imported_to_crm = 0

    return jsonify({
        'success': True,
        'stats': stats,
        'crm_imported': imported_to_crm,
        'requires_review': True,
        'download_file': output_filename,
        'job_id': uid,
        'skip_trace_available': SKIP_TRACE_PROVIDER != 'none'
    })


@app.route('/api/jobs/<job_id>')
def processing_job(job_id):
    meta = load_job_meta(job_id)
    if not meta:
        return jsonify({'error': 'Processing job not found'}), 404
    snapshot = processing_job_snapshot(meta)
    output_filename = snapshot['output_filename']
    snapshot['download_available'] = bool(
        output_filename
        and (
            get_crm().get_processing_artifact(filename=output_filename)
            or os.path.exists(
                os.path.join(
                    app.config['OUTPUT_FOLDER'],
                    os.path.basename(output_filename),
                )
            )
        )
    )
    return jsonify({'success': True, 'job': snapshot})


def _approved_import_snapshot(job_id):
    meta = load_job_meta(job_id)
    if not meta:
        raise FileNotFoundError('Processing job not found')
    verified_filename = meta.get('assessor_output_filename')
    if not verified_filename:
        raise ValueError('Assessor verification has not produced an approved workbook')
    workbook_path = materialize_artifact(
        job_id,
        'assessor',
        verified_filename,
        app.config['OUTPUT_FOLDER'],
    )
    if not workbook_path:
        raise FileNotFoundError('Assessor-verified workbook not found')
    leads = pd.read_excel(
        workbook_path,
        sheet_name='Prequalified - Verify',
        engine='openpyxl',
    )
    if 'Current Owner Verification' not in leads:
        raise ValueError('Current-owner decisions are missing')
    approved = leads[
        leads['Current Owner Verification'] == 'Verified candidate'
    ].copy()
    tax_id_column = find_column(approved, ['TAX', 'ID'])
    owner_column = (
        'Current Assessor Owner'
        if 'Current Assessor Owner' in approved
        else find_column(approved, ['OWNER', 'NAME'])
    )
    identity_parts = [
        f"{row.get(tax_id_column, '')}|{row.get(owner_column, '')}"
        for _, row in approved.iterrows()
    ]
    token_payload = (
        f"{job_id}|{verified_filename}|"
        + '|'.join(sorted(identity_parts))
    )
    approval_token = hashlib.sha256(token_payload.encode()).hexdigest()
    total_due_column = find_column(approved, ['TOTAL', 'DUE'])
    total_debt = (
        float(pd.to_numeric(approved[total_due_column], errors='coerce').fillna(0).sum())
        if total_due_column
        else 0
    )
    decision_counts = (
        leads['Current Owner Verification']
        .fillna('Not checked')
        .astype(str)
        .value_counts()
        .to_dict()
    )
    return {
        'meta': meta,
        'verified_filename': verified_filename,
        'approved': approved,
        'approval_token': approval_token,
        'total_debt': total_debt,
        'decision_counts': {
            str(key): int(value) for key, value in decision_counts.items()
        },
    }


@app.route('/api/import/<job_id>/preview')
def preview_approved_import(job_id):
    try:
        snapshot = _approved_import_snapshot(job_id)
    except FileNotFoundError as error:
        return jsonify({'error': str(error)}), 404
    except ValueError as error:
        return jsonify({'error': str(error)}), 409
    return jsonify(
        {
            'success': True,
            'job_id': job_id,
            'approved_candidates': len(snapshot['approved']),
            'total_debt': snapshot['total_debt'],
            'decision_counts': snapshot['decision_counts'],
            'approval_token': snapshot['approval_token'],
            'already_committed': bool(
                snapshot['meta'].get('crm_import_approved_at')
            ),
            'previously_imported': int(
                snapshot['meta'].get('crm_imported', 0)
            ),
        }
    )


@app.route('/api/import/<job_id>/commit', methods=['POST'])
def commit_approved_import(job_id):
    payload = request.get_json(silent=True) or {}
    try:
        snapshot = _approved_import_snapshot(job_id)
    except FileNotFoundError as error:
        return jsonify({'error': str(error)}), 404
    except ValueError as error:
        return jsonify({'error': str(error)}), 409
    if payload.get('approval_token') != snapshot['approval_token']:
        return jsonify(
            {'error': 'Approval preview is stale or was not confirmed'}
        ), 409
    if payload.get('confirmation') != 'IMPORT VERIFIED CANDIDATES':
        return jsonify({'error': 'Explicit import confirmation is required'}), 400

    approved = snapshot['approved']
    columns = {
        'tax_id': find_column(approved, ['TAX', 'ID']),
        'owner_name': (
            'Current Assessor Owner'
            if 'Current Assessor Owner' in approved
            else find_column(approved, ['OWNER', 'NAME'])
        ),
        'total_due': find_column(approved, ['TOTAL', 'DUE']),
        'phone': find_column(approved, ['PHONE']),
        'mailing_address': find_column(approved, ['ADDRESS']),
        'mailing_city': find_column(approved, ['OWNR_ADDR', '6']),
        'mailing_state': find_column(approved, ['OWNR_ADDR', 'ST']),
        'zip_code': find_column(approved, ['ZIP']),
        'street_number': find_column(approved, ['ST_NO']),
        'street_name': find_column(approved, ['ST_NAME']),
        'street_type': find_column(approved, ['ST_STREET', 'TYPE']),
        'property_city': find_column(approved, ['ST_CITY']),
        'deceased_flag': (
            'Deceased Evidence'
            if 'Deceased Evidence' in approved
            else None
        ),
        'mailing_signal': (
            'Absentee Signal' if 'Absentee Signal' in approved else None
        ),
    }
    required = ('tax_id', 'owner_name', 'total_due')
    if any(not columns[field] for field in required):
        return jsonify({'error': 'Required CRM import columns are missing'}), 409

    import_job = {
        'uid': job_id,
        'source_filename': snapshot['meta']['source_filename'],
        'output_filename': snapshot['verified_filename'],
        'tax_year': snapshot['meta']['tax_year'],
        'stats': {
            **snapshot['meta'].get('stats', {}),
            'approved_candidates': len(approved),
        },
    }
    imported = get_crm().import_leads(approved, import_job, columns)
    snapshot['meta']['crm_import_approved_at'] = datetime.now().isoformat()
    snapshot['meta']['crm_imported'] = int(
        snapshot['meta'].get('crm_imported', 0)
    ) + imported
    snapshot['meta']['crm_approval_token'] = snapshot['approval_token']
    snapshot['meta']['status'] = 'imported'
    save_job_meta(snapshot['meta'])
    return jsonify(
        {
            'success': True,
            'approved_candidates': len(approved),
            'imported': imported,
            'duplicates_skipped': len(approved) - imported,
            'crm_total': get_crm().dashboard_metrics()['actionable_leads'],
        }
    )


@app.route('/api/assessor/verify/<job_id>', methods=['POST'])
def verify_assessor_batch(job_id):
    meta = load_job_meta(job_id)
    if not meta:
        return jsonify({'error': 'Processing job not found'}), 404

    source_output_filename = meta.get(
        'assessor_output_filename',
        meta['output_filename'],
    )
    source_kind = (
        'assessor' if meta.get('assessor_output_filename')
        else 'qualification'
    )
    workbook_path = materialize_artifact(
        job_id,
        source_kind,
        source_output_filename,
        app.config['OUTPUT_FOLDER'],
    )
    if not workbook_path:
        return jsonify({'error': 'Qualification workbook not found'}), 404

    payload = request.get_json(silent=True) or {}
    configured_limit = int(app.config.get('ASSESSOR_BATCH_LIMIT', 25))
    try:
        requested_limit = int(payload.get('limit', configured_limit))
    except (TypeError, ValueError):
        return jsonify({'error': 'Batch limit must be a number'}), 400
    limit = min(max(requested_limit, 1), configured_limit)
    force = bool(payload.get('force', False))

    sheets = pd.read_excel(workbook_path, sheet_name=None, engine='openpyxl')
    leads = sheets.get('Prequalified - Verify')
    if leads is None:
        return jsonify({'error': 'Prequalified sheet not found'}), 409
    pid_column = find_column(leads, ['PID'])
    owner_column = find_column(leads, ['OWNER', 'NAME'])
    if not pid_column or not owner_column:
        return jsonify({'error': 'PID or owner column not found'}), 409

    repository = get_crm()
    client = app.config['ASSESSOR_CLIENT_FACTORY']()
    results = []
    processed = 0
    cache_hits = 0
    for index, row in leads.iterrows():
        if processed >= limit:
            break
        if str(row.get('Current Owner Verification', '')).strip() not in (
            '',
            'Not checked',
            'nan',
        ):
            continue
        account_no = normalize_account_no(row.get(pid_column))
        if not account_no:
            continue
        cached = None if force else repository.get_assessor_verification(account_no)
        if cached:
            cache_hits += 1
            fetched_at = cached['fetched_at']
            if hasattr(fetched_at, 'isoformat'):
                fetched_at = fetched_at.isoformat()
            result = AssessorResult(
                account_no=cached['account_no'],
                status=cached['status'],
                source_url=cached['source_url'],
                current_owner=cached['current_owner'] or '',
                account_type=cached['account_type'] or '',
                vacant=bool(cached['vacant']),
                error=cached['error'] or '',
                fetched_at=fetched_at,
            )
        else:
            result = client.fetch(row.get(pid_column))
            repository.save_assessor_verification(result)

        decision, reason = verification_decision(row.get(owner_column), result)
        leads.at[index, 'Current Owner Verification'] = decision
        leads.at[index, 'Current Assessor Owner'] = result.current_owner
        leads.at[index, 'Assessor Account Type'] = result.account_type
        leads.at[index, 'Assessor Vacant'] = 'Yes' if result.vacant else 'No'
        leads.at[index, 'Assessor Verification Reason'] = reason
        leads.at[index, 'Assessor Checked At'] = result.fetched_at
        leads.at[index, 'Assessor URL'] = result.source_url
        results.append(
            {
                **result.as_dict(),
                'source_owner': str(row.get(owner_column) or ''),
                'decision': decision,
                'decision_reason': reason,
            }
        )
        processed += 1

    sheets['Prequalified - Verify'] = leads
    verified_filename = f"Assessor_Verified_{meta['output_filename']}"
    verified_path = os.path.join(app.config['OUTPUT_FOLDER'], verified_filename)
    save_excel_formatted(sheets, verified_path)
    meta['assessor_output_filename'] = verified_filename
    meta['assessor_last_batch'] = {
        'processed': processed,
        'cache_hits': cache_hits,
        'completed_at': datetime.now().isoformat(),
    }

    counts = {}
    decisions = (
        leads['Current Owner Verification']
        .fillna('Not checked')
        .astype(str)
    )
    remaining = int(decisions.isin(('', 'Not checked', 'nan')).sum())
    checked = int(len(leads) - remaining)
    for decision, count in decisions.value_counts().to_dict().items():
        if decision not in ('', 'Not checked', 'nan'):
            counts[str(decision)] = int(count)
    meta['assessor_progress'] = {
        'checked': checked,
        'total': int(len(leads)),
        'remaining': remaining,
        'decision_counts': counts,
    }
    meta['status'] = (
        'ready_for_approval' if remaining == 0
        else 'assessor_in_progress'
    )
    save_job_meta(meta)
    persist_artifact(job_id, 'assessor', verified_filename, verified_path)
    return jsonify(
        {
            'success': True,
            'processed': processed,
            'cache_hits': cache_hits,
            'counts': counts,
            'results': results,
            'download_file': verified_filename,
            'remaining_estimate': remaining,
            'job': processing_job_snapshot(meta),
        }
    )


@app.route('/skiptrace/<job_id>', methods=['POST'])
def skiptrace(job_id):
    meta = load_job_meta(job_id)
    if not meta:
        return jsonify({'error': 'Job not found. Please process a list first.'}), 404

    source_filename = meta.get(
        'assessor_output_filename',
        meta['output_filename'],
    )
    source_kind = (
        'assessor' if meta.get('assessor_output_filename')
        else 'qualification'
    )
    clean_path = materialize_artifact(
        job_id,
        source_kind,
        source_filename,
        app.config['OUTPUT_FOLDER'],
    )
    if not clean_path:
        return jsonify({'error': 'Cleaned file not found.'}), 404

    df = pd.read_excel(
        clean_path,
        engine='openpyxl',
        sheet_name='Prequalified - Verify',
    )

    enriched_df, trace_stats = run_skip_tracing(df)

    if 'error' in trace_stats:
        return jsonify({'error': trace_stats['error']}), 500

    date_str = datetime.now().strftime('%Y%m%d')
    enriched_filename = f'Enriched_Leads_{meta["tax_year"]}_{date_str}_{job_id}.xlsx'
    enriched_path = os.path.join(app.config['OUTPUT_FOLDER'], enriched_filename)
    enriched_df.to_excel(enriched_path, index=False)
    meta['skiptrace_output_filename'] = enriched_filename
    persist_artifact(job_id, 'skiptrace', enriched_filename, enriched_path)
    save_job_meta(meta)

    return jsonify({
        'success': True,
        'trace_stats': trace_stats,
        'download_file': enriched_filename
    })


@app.route('/download/<filename>')
def download(filename):
    safe_name = os.path.basename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], safe_name)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=safe_name)
    artifact = get_crm().get_processing_artifact(filename=safe_name)
    if not artifact:
        return 'File not found', 404
    return send_file(
        io.BytesIO(artifact['content']),
        as_attachment=True,
        download_name=safe_name,
        mimetype='application/octet-stream',
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
