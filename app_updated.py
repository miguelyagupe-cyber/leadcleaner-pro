from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import re
import os
import uuid
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
app.config['OUTPUT_FOLDER'] = '/tmp/outputs'
os.makedirs('/tmp/uploads', exist_ok=True)
os.makedirs('/tmp/outputs', exist_ok=True)

BUSINESS_PATTERNS = [
    r'\bLLC\b', r'\bL\.L\.C\.?\b',
    r'\bINC\.?\b', r'\bINCORPORATED\b',
    r'\bCORP\.?\b', r'\bCORPORATION\b',
    r'\bLTD\.?\b', r'\bLIMITED\b',
    r'\bCOMPANY\b', r'\bCO\.\b',
    r'\bGROUP\b', r'\bENTERPRISES?\b',
    r'\bASSOCIATES?\b', r'\bPARTNERS?\b',
    r'\bHOLDINGS?\b', r'\bREALTY\b',
    r'\bREAL ESTATE\b', r'\bPROPERTIES\b',
    r'\bINVESTMENTS?\b', r'\bVENTURES?\b',
    r'\bCHURCH\b', r'\bCHAPEL\b', r'\bMINISTRIES?\b',
    r'\bFOUNDATION\b', r'\bCATTLE\b', r'\bRANCH\b',
    r'\bFARMS?\b', r'\bCULTIVATION\b',
]

COLUMN_RENAME = {
    'COMMENTS': 'Notes',
    'Total Due': 'Total Due ($)',
    'Tax ID': 'Tax ID',
    'PROPERTY OWNER NAME': 'Owner Name',
    'Phone': 'Phone',
    'OWNR_ADDR 2': 'Mailing Address',
    'OWNR_ADDR 3': 'Property Address',
    'OWNR_ADDR 6': 'City',
    'OWNR_ADDR ST': 'State',
    'ZIP': 'ZIP Code',
    'ST_NO': 'Street No.',
    'ST_Dir': 'Street Dir.',
    'ST_NAME': 'Street Name',
    'ST_STREET_TYPE': 'Street Type',
    'Legal Description': 'Legal Description',
    'SDName': 'School District',
    'TAX  YEAR': 'Tax Year',
    'SCHOOL': 'School Code',
    'ST_SUFFIX': 'Street Suffix',
}

COLUMN_ORDER = [
    'PROPERTY OWNER NAME', 'Total Due', 'Phone', 'Tax ID',
    'OWNR_ADDR 3', 'OWNR_ADDR 6', 'OWNR_ADDR ST', 'ZIP',
    'ST_NO', 'ST_Dir', 'ST_NAME', 'ST_STREET_TYPE',
    'OWNR_ADDR 2', 'Legal Description', 'SDName', 'TAX  YEAR',
    'SCHOOL', 'ST_SUFFIX', 'COMMENTS',
]

COLUMN_WIDTHS = {
    'Owner Name': 32, 'Total Due ($)': 14, 'Phone': 22, 'Tax ID': 12,
    'Property Address': 28, 'City': 16, 'State': 8, 'ZIP Code': 10,
    'Street No.': 10, 'Street Dir.': 10, 'Street Name': 18, 'Street Type': 12,
    'Mailing Address': 28, 'Legal Description': 35, 'School District': 20,
    'Tax Year': 10, 'School Code': 12, 'Street Suffix': 12, 'Notes': 40,
}

CENTER_COLS = {'Total Due ($)', 'Tax Year', 'Tax ID', 'ZIP Code', 'Street No.'}


def is_business(name):
    if pd.isna(name):
        return False
    n = str(name).upper()
    if re.search(r'\bTRUST(EE)?\b', n):
        return False
    return any(re.search(p, n) for p in BUSINESS_PATTERNS)


def clean_leads(df, tax_year):
    stats = {'original': len(df)}
    tax_col = next((c for c in df.columns if 'TAX' in c.upper() and 'YEAR' in c.upper()), None)
    if tax_col:
        df[tax_col] = pd.to_numeric(df[tax_col], errors='coerce')
        df = df[df[tax_col] == tax_year].copy()
    stats['after_year_filter'] = len(df)
    df['_is_business'] = df['PROPERTY OWNER NAME'].apply(is_business)
    df = df[~df['_is_business']].copy()
    df = df.drop(columns=['_is_business'])
    stats['after_business_filter'] = len(df)
    df = df.dropna(how='all')
    stats['final'] = len(df)
    if 'Total Due' in df.columns:
        df = df.sort_values('Total Due', ascending=False)
    stats['removed_year'] = stats['original'] - stats['after_year_filter']
    stats['removed_business'] = stats['after_year_filter'] - stats['after_business_filter']
    stats['with_phone'] = int(df['Phone'].notna().sum()) if 'Phone' in df.columns else 0
    stats['without_phone'] = stats['final'] - stats['with_phone']
    stats = {k: int(v) for k, v in stats.items()}
    return df, stats


def reorder_and_rename(df):
    existing = [c for c in COLUMN_ORDER if c in df.columns]
    remaining = [c for c in df.columns if c not in existing]
    df = df[existing + remaining]
    df = df.rename(columns=COLUMN_RENAME)
    return df


def format_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_fill = PatternFill('solid', start_color='1A1A1A', end_color='1A1A1A')
    header_font = Font(bold=True, color='C9A84C', name='Arial', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_fill_1 = PatternFill('solid', start_color='F9F9F9', end_color='F9F9F9')
    row_fill_2 = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
    data_font = Font(name='Arial', size=9)
    thin = Side(style='thin', color='E0E0E0')
    border = Border(bottom=thin)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        col_name = cell.value
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 15)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
        for col_idx, cell in enumerate(row, 1):
            cell.font = data_font
            cell.fill = fill
            cell.border = border
            col_name = ws.cell(row=1, column=col_idx).value
            if col_name in CENTER_COLS:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=False)
            if col_name == 'Total Due ($)':
                cell.number_format = '$#,##0.00'
    wb.save(path)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    tax_year = request.form.get('tax_year', '2022')
    try:
        tax_year = int(tax_year)
    except ValueError:
        return jsonify({'error': 'Invalid tax year'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv']:
        return jsonify({'error': 'File must be .xlsx, .xls or .csv'}), 400
    uid = str(uuid.uuid4())[:8]
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], f'{uid}_input{ext}')
    file.save(upload_path)
    try:
        df = pd.read_csv(upload_path) if ext == '.csv' else pd.read_excel(upload_path, engine='openpyxl')
    except Exception as e:
        return jsonify({'error': f'Could not read file: {str(e)}'}), 400
    try:
        cleaned_df, stats = clean_leads(df, tax_year)
        cleaned_df = reorder_and_rename(cleaned_df)
    except Exception as e:
        return jsonify({'error': f'Error during cleaning: {str(e)}'}), 500
    date_str = datetime.now().strftime('%Y%m%d')
    output_filename = f'Clean_Leads_{tax_year}_{date_str}_{uid}.xlsx'
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
    cleaned_df.to_excel(output_path, index=False, sheet_name='Clean Leads')
    try:
        format_excel(output_path)
    except Exception:
        pass
    return jsonify({'success': True, 'stats': stats, 'download_file': output_filename})


@app.route('/download/<filename>')
def download(filename):
    safe_name = os.path.basename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], safe_name)
    if not os.path.exists(filepath):
        return 'File not found', 404
    return send_file(filepath, as_attachment=True, download_name=safe_name)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
